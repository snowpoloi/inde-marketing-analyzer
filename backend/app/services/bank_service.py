from __future__ import annotations

import csv
import hashlib
import html
import io
import re
import unicodedata
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from html.parser import HTMLParser
from typing import Any

from dateutil import parser as date_parser

from app.services.parsing import dec_to_float


_NBG_ACCOUNT_LABEL = "".join(
    chr(value)
    for value in (945, 961, 953, 952, 956, 959, 963, 32, 955, 959, 947, 945, 961, 953, 945, 963, 956, 959, 965)
)


class _TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag == "tr":
            self._row = []
        elif tag in {"td", "th"} and self._row is not None:
            self._cell = []

    def handle_data(self, data: str) -> None:
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag in {"td", "th"} and self._row is not None and self._cell is not None:
            self._row.append(html.unescape("".join(self._cell)).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if any(_clean_text(cell) for cell in self._row):
                self.rows.append(self._row)
            self._row = None


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("_x000D_", " ").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def _normalise_text(value: Any) -> str:
    text = _clean_text(value).casefold()
    return "".join(char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn")


def _money(value: Any, *, cents: bool = False) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        parsed = value
    elif isinstance(value, int):
        parsed = Decimal(value)
    elif isinstance(value, float):
        parsed = Decimal(str(value))
    else:
        text = _clean_text(value)
        text = text.replace("\u20ac", "").replace("EUR", "").replace(" ", "").replace("\u00a0", "")
        if not text:
            return Decimal("0")
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")
        try:
            parsed = Decimal(text)
        except (InvalidOperation, ValueError):
            return Decimal("0")
    if cents:
        parsed = parsed / Decimal("100")
    return parsed.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def _parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date_parser.parse(str(value), dayfirst=True).date()
    except (TypeError, ValueError, OverflowError):
        return None


def _csv_rows(content: bytes) -> list[list[str]]:
    for encoding in ("utf-8-sig", "cp1253", "windows-1253"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode("utf-8", errors="ignore")

    try:
        dialect = csv.Sniffer().sniff(text[:2000], delimiters=";,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"
    return list(csv.reader(io.StringIO(text), dialect))


def _html_rows(content: bytes) -> list[list[str]]:
    parser = _TableParser()
    parser.feed(content.decode("utf-8", errors="ignore"))
    return parser.rows


def _xlsx_rows(content: bytes) -> tuple[list[list[Any]], dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    metadata: dict[str, Any] = {}
    sheets: list[list[list[Any]]] = []
    for sheet in workbook.worksheets:
        sheet_rows: list[list[Any]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            row = list(raw_row)
            sheet_rows.append(row)
            joined = " ".join(_clean_text(cell) for cell in row)
            for index, cell in enumerate(row):
                if _NBG_ACCOUNT_LABEL not in _normalise_text(cell):
                    continue
                account = next((_clean_text(value) for value in row[index + 1 :] if _clean_text(value)), "")
                if account:
                    metadata["account"] = account
            iban = _extract_iban(joined)
            if iban and not metadata.get("account"):
                metadata["account"] = iban
        sheets.append(sheet_rows)

    for rows in sheets:
        if _find_nbg_header(rows) is not None:
            return rows, metadata
    for rows in sheets:
        if _find_piraeus_header(rows) is not None:
            return rows, metadata
    return (sheets[0] if sheets else []), metadata


def _extract_iban(value: str) -> str | None:
    match = re.search(r"\bGR[0-9A-Z ]{12,34}\b", value.replace("\u00a0", " "))
    if not match:
        return None
    return re.sub(r"\s+", "", match.group(0))


def _data_rows(rows: list[list[Any]], start_index: int, mapping: dict[str, int]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for values in rows[start_index + 1 :]:
        if not any(_clean_text(value) for value in values):
            continue
        record = {field: values[index] if index < len(values) else None for field, index in mapping.items()}
        if _parse_date(record.get("transaction_date")) is not None:
            records.append(record)
    return records


def _find_nbg_header(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows):
        if any(_clean_text(cell).casefold() == "valeur" for cell in row):
            return index
    return None


def _find_piraeus_header(rows: list[list[Any]]) -> int | None:
    for index in range(len(rows) - 1):
        if len(rows[index]) >= 8 and len(rows[index + 1]) >= 8 and _parse_date(rows[index + 1][2]):
            return index
    return None


def _find_wallet_header(rows: list[list[Any]]) -> int | None:
    for index in range(len(rows) - 1):
        if len(rows[index]) >= 8 and len(rows[index + 1]) >= 8 and _parse_date(rows[index + 1][1]):
            return index
    return None


def parse_bank_export(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = filename.lower().rsplit(".", 1)[-1]
    metadata: dict[str, Any] = {}
    bank_name = "Bank"
    cents = False
    mapping: dict[str, int]

    if suffix == "csv":
        rows = _csv_rows(content)
        if len(rows) < 2:
            return []
        bank_name = "Bank CSV"
        mapping = {
            "transaction_date": 0,
            "value_date": 1,
            "description": 2,
            "amount": 3,
            "balance": 4,
            "counterparty": 5,
        }
        records = _data_rows(rows, 0, mapping)
    elif suffix == "xls" and content.lstrip().lower().startswith(b"<!doctype html"):
        rows = _html_rows(content)
        header_index = _find_wallet_header(rows)
        if header_index is None:
            return []
        bank_name = "Viva Wallet"
        mapping = {
            "reference": 0,
            "transaction_date": 1,
            "value_date": 3,
            "description": 4,
            "counterparty": 5,
            "currency": 6,
            "amount": 7,
            "balance": 8,
        }
        records = _data_rows(rows, header_index, mapping)
    else:
        rows, metadata = _xlsx_rows(content)
        nbg_header = _find_nbg_header(rows)
        if nbg_header is not None:
            bank_name = "NBG"
            mapping = {
                "reference": 0,
                "transaction_date": 1,
                "value_date": 3,
                "category_raw": 5,
                "operation_raw": 6,
                "amount": 8,
                "currency": 9,
                "debit_credit": 10,
                "description": 11,
                "balance": 12,
                "counterparty": 14,
                "counterparty_account": 15,
                "counterparty_bank": 16,
            }
            records = _data_rows(rows, nbg_header, mapping)
        else:
            piraeus_header = _find_piraeus_header(rows)
            if piraeus_header is None:
                return []
            bank_name = "Piraeus"
            mapping = {
                "category_raw": 0,
                "description": 1,
                "transaction_date": 2,
                "value_date": 3,
                "reference": 4,
                "amount": 5,
                "currency": 6,
                "balance": 7,
            }
            records = _data_rows(rows, piraeus_header, mapping)

    parsed = [
        _normalise_record(filename, bank_name, record, metadata, cents=cents)
        for record in records
    ]
    return [record for record in parsed if record is not None]


def _normalise_record(
    filename: str,
    bank_name: str,
    record: dict[str, Any],
    metadata: dict[str, Any],
    *,
    cents: bool = False,
) -> dict[str, Any] | None:
    transaction_date = _parse_date(record.get("transaction_date"))
    if transaction_date is None:
        return None

    description = _clean_text(record.get("description")) or "-"
    reference = _clean_text(record.get("reference")) or None
    counterparty = _clean_text(record.get("counterparty")) or None
    amount = _money(record.get("amount"), cents=cents)
    debit_credit = _normalise_text(record.get("debit_credit"))
    if debit_credit.startswith("\u03c7") and amount > 0:
        amount = -amount

    category_raw = _clean_text(record.get("category_raw"))
    operation_raw = _clean_text(record.get("operation_raw"))
    detail_bits = [bit for bit in (category_raw, operation_raw, reference) if bit and bit not in description]
    if detail_bits:
        description = f"{description} | {' | '.join(detail_bits)}"

    balance = _money(record.get("balance"), cents=cents)
    currency = _clean_text(record.get("currency")) or "EUR"
    debit = abs(amount) if amount < 0 else Decimal("0")
    credit = amount if amount > 0 else Decimal("0")
    account = metadata.get("account") or _clean_text(record.get("counterparty_account")) or None
    category = _categorise(description, counterparty, amount)
    transaction_type = "deposit" if amount > 0 else "expense"

    source_basis = "|".join(
        [
            bank_name,
            account or "",
            transaction_date.isoformat(),
            str(amount),
            description,
            counterparty or "",
            reference or "",
            str(balance),
        ]
    )
    return {
        "bank_name": bank_name,
        "account": account,
        "transaction_date": transaction_date,
        "value_date": _parse_date(record.get("value_date")),
        "description": description,
        "counterparty": counterparty,
        "reference": reference,
        "amount": amount,
        "debit": debit,
        "credit": credit,
        "balance": balance,
        "currency": currency,
        "category": category,
        "transaction_type": transaction_type,
        "source_filename": filename,
        "source_key": hashlib.sha256(source_basis.encode("utf-8")).hexdigest(),
        "raw": {key: _clean_text(value) for key, value in record.items()},
    }


def _contains_any(text: str, tokens: tuple[str, ...]) -> bool:
    return any(token in text for token in tokens)


def _categorise(description: str, counterparty: str | None, amount: Decimal) -> str:
    text = _normalise_text(f"{description} {counterparty or ''}")
    if amount > 0:
        if _contains_any(
            text,
            (
                "\u03c0\u03b1\u03c1\u03b1\u03b3\u03b3\u03b5\u03bb",
                "\u03c0\u03bb\u03b7\u03c1\u03c9\u03bc",
                "\u03b5\u03be\u03bf\u03c6\u03bb",
                "order",
            ),
        ):
            return "Customer deposits"
        if _contains_any(text, ("\u03b5\u03ba\u03ba\u03b1\u03b8\u03b1\u03c1", "settlement", "wallet")):
            return "Card settlements"
        return "Deposits"

    if _contains_any(text, ("\u03c0\u03c1\u03bf\u03bc\u03b7\u03b8\u03b5\u03b9", "commission", "comm.", "bank fee")):
        return "Bank fees"
    if _contains_any(text, ("\u03c7\u03c1\u03b5\u03c9\u03c3",)) and "iban" in text:
        return "Bank fees"
    if _contains_any(text, ("google", "meta", "facebook", "ads", "tiktok")):
        return "Advertising"
    if _contains_any(text, ("\u03bc\u03b9\u03c3\u03b8\u03bf\u03b4\u03bf\u03c3\u03b9", "payroll")):
        return "Payroll"
    if _contains_any(text, ("\u03ba\u03bf\u03c5\u03c1\u03b9\u03b5\u03c1", "courier", "\u03bc\u03b5\u03c4\u03b1\u03c6\u03bf\u03c1", "\u03b1\u03c0\u03bf\u03c3\u03c4\u03bf\u03bb")):
        return "Shipping / logistics"
    if _contains_any(text, ("\u03c6\u03bf\u03c1\u03bf", "tax", "\u03b1\u03b1\u03b4\u03b5", "efka", "\u03b5\u03c6\u03ba\u03b1")):
        return "Taxes / social insurance"
    if _contains_any(text, ("inde interior deals", "inde inde")):
        return "Internal transfer"
    if _contains_any(text, ("\u03b1\u03b3\u03bf\u03c1\u03b1", "card purchase")) and _contains_any(text, ("\u03ba\u03b1\u03c1\u03c4", "card")):
        return "Card purchases"
    if _contains_any(text, ("\u03c4\u03b9\u03bc\u03bf\u03bb\u03bf\u03b3", "\u03c0\u03c1\u03bf\u03bc\u03b7\u03b8\u03b5\u03c5", "\u03b5\u03bc\u03b2\u03b1\u03c3\u03bc", "invoice", "supplier")):
        return "Supplier payments"
    return "Suppliers / other expenses"


def import_bank_transactions(db: Any, filename: str, content: bytes) -> dict[str, int | str]:
    from sqlalchemy import select

    from app.models import BankTransaction

    parsed = parse_bank_export(filename, content)
    if not parsed:
        return {"filename": filename, "total": 0, "imported": 0, "skipped": 0}

    keys = [row["source_key"] for row in parsed]
    existing = set(db.scalars(select(BankTransaction.source_key).where(BankTransaction.source_key.in_(keys))).all())
    nbg_rows = [row for row in parsed if row["bank_name"] == "NBG"]
    legacy_nbg: dict[tuple[Any, ...], list[Any]] = {}
    if nbg_rows:
        first_date = min(row["transaction_date"] for row in nbg_rows)
        last_date = max(row["transaction_date"] for row in nbg_rows)
        for transaction in db.scalars(
            select(BankTransaction).where(
                BankTransaction.bank_name == "NBG",
                BankTransaction.transaction_date >= first_date,
                BankTransaction.transaction_date <= last_date,
            )
        ).all():
            fingerprint = _nbg_transaction_fingerprint(transaction)
            if fingerprint:
                legacy_nbg.setdefault(fingerprint, []).append(transaction)

    imported = 0
    reconciled = 0
    for row in parsed:
        if row["source_key"] in existing:
            continue
        fingerprint = _nbg_record_fingerprint(row)
        matching_legacy = legacy_nbg.get(fingerprint, []) if fingerprint else []
        if matching_legacy:
            for transaction in matching_legacy:
                transaction.account = row["account"]
            reconciled += len(matching_legacy)
            continue
        db.add(BankTransaction(**row))
        imported += 1
    db.commit()
    return {
        "filename": filename,
        "total": len(parsed),
        "imported": imported,
        "reconciled": reconciled,
        "skipped": len(parsed) - imported - reconciled,
    }


def bank_dashboard(db: Any, date_from: date, date_to: date) -> dict[str, Any]:
    from sqlalchemy import func, select
    from sqlalchemy.orm import selectinload

    from app.models import BankTransaction, OpenCartOrder

    transactions = db.scalars(
        select(BankTransaction)
        .where(BankTransaction.transaction_date >= date_from, BankTransaction.transaction_date <= date_to)
        .order_by(BankTransaction.transaction_date.desc(), BankTransaction.created_at.desc())
    ).all()

    orders = db.scalars(
        select(OpenCartOrder)
        .options(selectinload(OpenCartOrder.products))
        .where(func.date(OpenCartOrder.date_added).between(date_from - timedelta(days=14), date_to + timedelta(days=14)))
    ).all()
    transactions, duplicate_transactions = _unique_bank_transactions(transactions)
    category_totals: dict[str, dict[str, Any]] = {}
    bank_totals: dict[tuple[str, str], dict[str, Any]] = {}
    deposit_rows: list[dict[str, Any]] = []
    transaction_rows: list[dict[str, Any]] = []
    deposit_matches = match_bank_deposits_for_orders(transactions, orders)
    order_lookup = {order.order_id: order for order in orders}
    orders_by_amount: dict[int, list[Any]] = {}
    for order in orders:
        orders_by_amount.setdefault(_money_key(order.total), []).append(order)

    for transaction in transactions:
        bank_bucket = bank_totals.setdefault(
            (transaction.bank_name, transaction.account or "-"),
            {
                "bank_name": transaction.bank_name,
                "account": transaction.account or "-",
                "transactions": 0,
                "credits": Decimal("0"),
                "debits": Decimal("0"),
            },
        )
        bank_bucket["transactions"] += 1
        if transaction.amount > 0:
            bank_bucket["credits"] += transaction.amount
        elif transaction.amount < 0:
            bank_bucket["debits"] += abs(transaction.amount)

        match = deposit_matches.get(transaction.id) if transaction.amount > 0 else None
        review_reason = (
            _deposit_review_reason(transaction, match, order_lookup, orders_by_amount)
            if transaction.amount > 0
            else None
        )
        row = {**_transaction_row(transaction), "match": match, "review_reason": review_reason}
        transaction_rows.append(row)
        if transaction.amount > 0:
            deposit_rows.append(row)
        elif transaction.amount < 0:
            bucket = category_totals.setdefault(
                transaction.category,
                {"category": transaction.category, "transactions": 0, "amount": Decimal("0")},
            )
            bucket["transactions"] += 1
            bucket["amount"] += abs(transaction.amount)

    deposit_total = sum((transaction.amount for transaction in transactions if transaction.amount > 0), Decimal("0"))
    expense_total = sum((abs(transaction.amount) for transaction in transactions if transaction.amount < 0), Decimal("0"))
    bank_fees = sum((abs(transaction.amount) for transaction in transactions if transaction.category == "Bank fees"), Decimal("0"))
    unmatched_deposits = sum(1 for row in deposit_rows if row["match"] is None)

    return {
        "summary": {
            "deposits": dec_to_float(deposit_total),
            "expenses": dec_to_float(expense_total),
            "net_cashflow": dec_to_float(deposit_total - expense_total),
            "bank_fees": dec_to_float(bank_fees),
            "transactions": len(transactions),
            "unmatched_deposits": unmatched_deposits,
            "ignored_duplicates": len(duplicate_transactions),
            "ignored_duplicate_amount": dec_to_float(
                sum((abs(transaction.amount) for transaction in duplicate_transactions), Decimal("0"))
            ),
        },
        "expense_categories": [
            {
                "category": row["category"],
                "transactions": row["transactions"],
                "amount": dec_to_float(row["amount"]),
            }
            for row in sorted(category_totals.values(), key=lambda item: item["amount"], reverse=True)
        ],
        "bank_totals": [
            {
                "bank_name": row["bank_name"],
                "account": row["account"],
                "transactions": row["transactions"],
                "credits": dec_to_float(row["credits"]),
                "debits": dec_to_float(row["debits"]),
                "net_cashflow": dec_to_float(row["credits"] - row["debits"]),
            }
            for row in sorted(bank_totals.values(), key=lambda item: (item["bank_name"].casefold(), item["account"]))
        ],
        "deposits": deposit_rows,
        "transactions": transaction_rows,
    }


def _unique_bank_transactions(transactions: list[Any]) -> tuple[list[Any], list[Any]]:
    """Exclude only exact cross-export duplicates from reporting totals.

    Import-level source keys already reject identical rows. This second check catches
    the same movement imported through two export formats without guessing from
    description text or amount alone.
    """
    seen: set[tuple[Any, ...]] = set()
    unique: list[Any] = []
    duplicates: list[Any] = []

    for transaction in transactions:
        fingerprint = _bank_transaction_fingerprint(transaction)
        if fingerprint and fingerprint in seen:
            duplicates.append(transaction)
            continue
        if fingerprint:
            seen.add(fingerprint)
        unique.append(transaction)
    return unique, duplicates


def _bank_transaction_fingerprint(transaction: Any) -> tuple[Any, ...] | None:
    nbg_fingerprint = _nbg_transaction_fingerprint(transaction)
    if nbg_fingerprint:
        return nbg_fingerprint

    reference = re.sub(r"[^a-z0-9]", "", _normalise_text(transaction.reference))
    if len(reference) >= 8:
        return ("reference", transaction.transaction_date.isoformat(), _money_key(transaction.amount), reference)

    account = _normalise_text(transaction.account)
    if account and transaction.balance is not None:
        return (
            "account_balance",
            account,
            transaction.transaction_date.isoformat(),
            _money_key(transaction.amount),
            _money_key(transaction.balance),
        )
    return None


def _nbg_record_fingerprint(record: dict[str, Any]) -> tuple[Any, ...] | None:
    if record.get("bank_name") != "NBG":
        return None
    return _nbg_fingerprint_values(
        record.get("transaction_date"),
        record.get("amount"),
        record.get("balance"),
        record.get("description"),
        record.get("counterparty"),
        record.get("reference"),
    )


def _nbg_transaction_fingerprint(transaction: Any) -> tuple[Any, ...] | None:
    if transaction.bank_name != "NBG":
        return None
    return _nbg_fingerprint_values(
        transaction.transaction_date,
        transaction.amount,
        transaction.balance,
        transaction.description,
        transaction.counterparty,
        transaction.reference,
    )


def _nbg_fingerprint_values(
    transaction_date: Any,
    amount: Any,
    balance: Any,
    description: Any,
    counterparty: Any,
    reference: Any,
) -> tuple[Any, ...] | None:
    if transaction_date is None or balance is None:
        return None
    return (
        "nbg",
        transaction_date.isoformat(),
        _money_key(amount),
        _money_key(balance),
        _normalise_text(description),
        _normalise_text(counterparty),
        _normalise_text(reference),
    )


def match_bank_deposits_for_orders(transactions: list[Any], orders: list[Any]) -> dict[Any, dict[str, Any] | None]:
    order_lookup = {order.order_id: order for order in orders}
    orders_by_amount: dict[int, list[Any]] = {}
    for order in orders:
        orders_by_amount.setdefault(_money_key(order.total), []).append(order)

    deposit_matches: dict[Any, dict[str, Any] | None] = {}
    deposit_transactions = [transaction for transaction in transactions if transaction.amount > 0]
    for transaction in deposit_transactions:
        deposit_matches[transaction.id] = _match_deposit(transaction, order_lookup, orders_by_amount)

    _enrich_split_deposit_matches(deposit_transactions, deposit_matches, order_lookup)
    return deposit_matches


def _transaction_row(transaction: BankTransaction) -> dict[str, Any]:
    return {
        "id": str(transaction.id),
        "bank_name": transaction.bank_name,
        "account": transaction.account,
        "transaction_date": transaction.transaction_date.isoformat(),
        "value_date": transaction.value_date.isoformat() if transaction.value_date else None,
        "description": transaction.description,
        "counterparty": transaction.counterparty,
        "reference": transaction.reference,
        "amount": dec_to_float(transaction.amount),
        "debit": dec_to_float(transaction.debit),
        "credit": dec_to_float(transaction.credit),
        "balance": dec_to_float(transaction.balance),
        "currency": transaction.currency or "EUR",
        "category": transaction.category,
        "transaction_type": transaction.transaction_type,
        "source_filename": transaction.source_filename,
    }


def _money_key(value: Any) -> int:
    return int((Decimal(str(value or 0)) * 100).to_integral_value(rounding=ROUND_HALF_UP))


def _extract_order_ids(text: str) -> list[str]:
    ids: list[str] = []
    normalised = _normalise_text(text)
    for pattern in (
        r"(?:\u03c0\u03b1\u03c1\u03b1\u03b3\u03b3\u03b5\u03bb\w*|order|ord)\D{0,20}(\d{4,8})",
        r"\b(1[0-9]{5}|[2-9][0-9]{4,7})\b",
    ):
        for match in re.finditer(pattern, normalised):
            value = match.group(1)
            if value not in ids and not value.startswith("2026"):
                ids.append(value)
    return ids


def _order_age_days(transaction: Any, order: Any) -> int:
    return (transaction.transaction_date - order.date_added.date()).days


def _is_after_order(transaction: Any, order: Any, *, max_days: int = 14) -> bool:
    age_days = _order_age_days(transaction, order)
    return 0 <= age_days <= max_days


def _order_text_score(transaction: Any, order: Any) -> int:
    text = _normalise_text(" ".join([transaction.description, transaction.counterparty or "", transaction.reference or ""]))
    raw = order.raw or {}
    values = [
        order.order_id,
        raw.get("firstname"),
        raw.get("lastname"),
        raw.get("customer"),
        raw.get("customer_name"),
        raw.get("payment_firstname"),
        raw.get("payment_lastname"),
        raw.get("shipping_firstname"),
        raw.get("shipping_lastname"),
        raw.get("email"),
        raw.get("telephone"),
    ]
    score = 0
    for value in values:
        token = _normalise_text(value)
        if len(token) >= 4 and token in text:
            score += 1
    return score


def _single_text_supported_candidate(transaction: Any, candidates: list[Any]) -> Any | None:
    scored = [(candidate, _order_text_score(transaction, candidate)) for candidate in candidates]
    supported = [candidate for candidate, score in scored if score > 0]
    if len(supported) == 1:
        return supported[0]
    return None


def _match_deposit(
    transaction: Any,
    order_lookup: dict[str, Any],
    orders_by_amount: dict[int, list[Any]],
) -> dict[str, Any] | None:
    text = " ".join([transaction.description, transaction.counterparty or "", transaction.reference or ""])
    for order_id in _extract_order_ids(text):
        order = order_lookup.get(order_id)
        if order and _is_after_order(transaction, order):
            return _order_match_payload(order, transaction.amount, "order reference")

    candidates = orders_by_amount.get(_money_key(transaction.amount), [])
    near = [order for order in candidates if _is_after_order(transaction, order)]
    if len(near) == 1:
        return _order_match_payload(near[0], transaction.amount, "same amount near date")
    if len(near) > 1:
        supported = _single_text_supported_candidate(transaction, near)
        if supported:
            return _order_match_payload(supported, transaction.amount, "same amount with depositor/comment match")
    return None


def _deposit_review_reason(
    transaction: Any,
    match: dict[str, Any] | None,
    order_lookup: dict[str, Any],
    orders_by_amount: dict[int, list[Any]],
) -> str | None:
    if transaction.amount <= 0 or match:
        return None

    text = " ".join([transaction.description, transaction.counterparty or "", transaction.reference or ""])
    referenced_ids = _extract_order_ids(text)
    for order_id in referenced_ids:
        order = order_lookup.get(order_id)
        if not order:
            return "Order reference not found in this order window."
        if not _is_after_order(transaction, order):
            return "Payment date is before the order date or too far after it."

    same_amount = [order for order in orders_by_amount.get(_money_key(transaction.amount), []) if _is_after_order(transaction, order)]
    if len(same_amount) > 1:
        return "Multiple orders have the same amount. Check depositor name and comments before matching."
    if len(same_amount) == 1:
        return "Same amount found, but no clear order reference or depositor/comment support."
    return "No safe order match found."


def _enrich_split_deposit_matches(
    transactions: list[Any],
    matches: dict[Any, dict[str, Any] | None],
    order_lookup: dict[str, Any],
) -> None:
    used_transaction_ids: set[Any] = set()

    for transaction in transactions:
        match = matches.get(transaction.id)
        if not match or match.get("amount_gap", 0) >= 0:
            continue

        order = order_lookup.get(match["order_id"])
        if not order:
            continue
        if not _is_after_order(transaction, order):
            continue

        total = Decimal(str(order.total or 0))
        residual = total - Decimal(str(transaction.amount or 0))
        if residual <= 0:
            continue

        candidates = [
            candidate
            for candidate in transactions
            if candidate.id != transaction.id
            and candidate.id not in used_transaction_ids
            and matches.get(candidate.id) is None
            and _money_key(candidate.amount) == _money_key(residual)
            and _is_after_order(candidate, order)
            and candidate.transaction_date <= transaction.transaction_date
            and (transaction.transaction_date - candidate.transaction_date).days <= 14
        ]
        if len(candidates) != 1:
            continue

        candidate = candidates[0]
        used_transaction_ids.add(candidate.id)
        matched_total = Decimal(str(transaction.amount or 0)) + Decimal(str(candidate.amount or 0))
        transaction_related = [_related_transaction_payload(candidate)]
        candidate_related = [_related_transaction_payload(transaction)]

        matches[transaction.id] = _order_match_payload(
            order,
            transaction.amount,
            "split payment order reference",
            matched_bank_total=matched_total,
            coverage_override="split",
            related_transactions=transaction_related,
        )
        matches[candidate.id] = _order_match_payload(
            order,
            candidate.amount,
            "split payment residual",
            matched_bank_total=matched_total,
            coverage_override=_residual_coverage(order, candidate.amount),
            related_transactions=candidate_related,
        )


def _related_transaction_payload(transaction: Any) -> dict[str, Any]:
    return {
        "id": str(transaction.id),
        "transaction_date": transaction.transaction_date.isoformat(),
        "bank_name": transaction.bank_name,
        "amount": dec_to_float(transaction.amount),
        "description": transaction.description,
        "reference": transaction.reference,
    }


def _residual_coverage(order: Any, amount: Decimal) -> str:
    total = Decimal(str(order.total or 0))
    shipping = Decimal(str(order.shipping or 0))
    product_amount = total - shipping
    if _money_key(amount) == _money_key(product_amount):
        return "products"
    if _money_key(amount) == _money_key(shipping):
        return "shipping"
    return "remaining"


def _order_match_payload(
    order: Any,
    amount: Decimal,
    reason: str,
    *,
    matched_bank_total: Decimal | None = None,
    coverage_override: str | None = None,
    related_transactions: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    total = Decimal(str(order.total or 0))
    shipping = Decimal(str(order.shipping or 0))
    product_amount = total - shipping
    comparison_amount = matched_bank_total or amount
    if coverage_override:
        coverage = coverage_override
    elif _money_key(amount) == _money_key(total):
        coverage = "all"
    elif _money_key(amount) == _money_key(product_amount):
        coverage = "products"
    elif _money_key(amount) == _money_key(shipping):
        coverage = "shipping"
    elif amount > total:
        coverage = "overpaid"
    else:
        coverage = "partial"
    return {
        "order_id": order.order_id,
        "status": order.order_status,
        "date_added": order.date_added.isoformat(),
        "order_total": dec_to_float(total),
        "product_amount": dec_to_float(product_amount),
        "shipping": dec_to_float(shipping),
        "matched_bank_total": dec_to_float(comparison_amount),
        "amount_gap": dec_to_float(comparison_amount - total),
        "payment_coverage": coverage,
        "match_reason": reason,
        "related_transactions": related_transactions or [],
    }
